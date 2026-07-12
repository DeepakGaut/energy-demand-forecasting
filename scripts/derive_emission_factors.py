"""
Derive generation-weighted emission factors (kg CO2e / kWh) per fuel type
from CO2_Database_V_21.0.xlsx ("Data" sheet).

Why generation-weighted, not a simple average:
    A simple mean across plant-units would let a tiny 5 MW diesel set
    count as much as a 4000 MW coal station. Weighting by each unit's
    2024-25 net generation means the result reflects the actual grid
    mix, which is what we need for CI.

Method:
    EF_fuel = Σ(Absolute Emissions t CO2) / Σ(Net Generation GWh × 1000)
    (t CO2 / MWh is numerically identical to kg CO2 / kWh — no
    conversion needed beyond GWh -> MWh.)

Run standalone to reproduce/verify the numbers hardcoded into
clean_and_compute_ci.py:
    python scripts/derive_emission_factors.py --input CO2_Database_V_21.0.xlsx
"""

import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl

# CEA "FUEL 1" categories -> our internal fuel buckets.
# COAL/GAS map 1:1. Everything else thermal (lignite, oil, naphtha,
# diesel) is small-volume and gets grouped into "others" to match the
# others_mwh column in nr.csv/wr.csv/etc.
FUEL_MAP = {
    "COAL": "coal",
    "GAS": "gas",
    "LIGN": "others",
    "OIL": "others",
    "NAPT": "others",
    "DISL": "others",
}

# Hydro, nuclear, wind, solar aren't meaningfully populated in this
# thermal-plant-focused database (hydro/nuclear rows exist with 0
# emissions; wind/solar don't appear at all). Standard zero-direct-
# emissions assumption applies to all four.
ZERO_EF_FUELS = ["hydro", "nuclear", "wind", "solar"]


def derive_emission_factors(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Data"]
    rows = ws.iter_rows(min_row=2, values_only=True)

    header = next(iter(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    col = {name: i for i, name in enumerate(header) if name}

    fuel1_idx = col.get("FUEL 1")
    gen_idx = next((i for n, i in col.items() if "Net" in str(n) and "Generation" in str(n)), None)
    emis_idx = next((i for n, i in col.items() if "Absolute" in str(n) and "Emissions" in str(n)), None)

    if fuel1_idx is None or gen_idx is None or emis_idx is None:
        raise ValueError(
            f"Could not find expected columns in 'Data' sheet. "
            f"Found headers: {list(col.keys())}. "
            f"Looking for 'FUEL 1', a generation column containing "
            f"'Net'+'Generation', and an emissions column containing "
            f"'Absolute'+'Emissions'."
        )

    gen_by_bucket = defaultdict(float)   # GWh
    emis_by_bucket = defaultdict(float)  # t CO2

    for row in rows:
        fuel1 = row[fuel1_idx]
        bucket = FUEL_MAP.get(fuel1)
        if bucket is None:
            continue
        gen_by_bucket[bucket] += row[gen_idx] or 0
        emis_by_bucket[bucket] += row[emis_idx] or 0

    factors = {}
    for bucket, gwh in gen_by_bucket.items():
        mwh = gwh * 1000
        factors[bucket] = (emis_by_bucket[bucket] / mwh) if mwh else 0.0

    for fuel in ZERO_EF_FUELS:
        factors[fuel] = 0.0

    return factors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=Path("CO2_Database_V_21.0.xlsx"))
    args = parser.parse_args()

    factors = derive_emission_factors(args.input)
    print("Derived emission factors (kg CO2e / kWh):")
    for fuel, ef in sorted(factors.items()):
        print(f"  {fuel:8s}: {ef:.4f}")


if __name__ == "__main__":
    main()